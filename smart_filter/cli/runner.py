from __future__ import annotations

import argparse
import json
import sys
import tempfile
from pathlib import Path
from typing import Sequence

from cli_core import EXIT_ERROR, EXIT_OK, create_base_parser, parse_cli_options
from json_contract_core import create_result_contract, validate_contract, write_json_file
from logging_core import create_logger
from tool_runtime_core import create_runtime_context

from smart_filter.app_info import APP_NAME, APP_VERSION
from smart_filter.domain.search_config import (
    ALL_FILE_TYPE_OPTION,
    ANALYSIS_MODE_FILE,
    ANALYSIS_MODE_FOLDER,
    DEFAULT_CATEGORY_NAME,
    DEFAULT_FILE_TYPE_OPTION,
    DEFAULT_SEARCH_SCOPE_OPTION,
    get_search_file_type_options,
    get_search_scope_options,
)
from smart_filter.domain.search_form_state import SearchFormState
from smart_filter.engine.search_engine import run_search
from smart_filter.output.result_contract import write_search_results_contract
from smart_filter.paths import (
    FACTORY_CATEGORIES_PATH,
    FACTORY_SETTINGS_PATH,
    OUTPUT_DIR,
    PROJECT_ROOT,
    ensure_project_directories,
)
from smart_filter.services.category_service import get_category_names, load_category_result
from smart_filter.services.settings_service import load_settings_result


def _comma_items(value: str | None) -> list[str]:
    if not value:
        return []
    return [item.strip() for item in str(value).split(",") if item.strip()]


def build_parser() -> argparse.ArgumentParser:
    parser = create_base_parser(
        tool_name="SmartFilter",
        description="Smart Filter CLI: búsqueda inteligente de archivos con salida JSON estándar.",
        version=APP_VERSION,
    )

    source = parser.add_mutually_exclusive_group()
    source.add_argument("--folder", type=Path, help="Carpeta a escanear.")
    source.add_argument("--file", type=Path, help="Archivo individual a analizar.")

    parser.add_argument("--query", "--search-text", dest="search_text", default="", help="Palabra o frase a buscar.")
    parser.add_argument("--context", "--must-contain", dest="context_filter", default="", help="Contexto requerido adicional separado por coma/;/. Ej: ShadowBackup, SmartFilter.")
    parser.add_argument("--category", default=DEFAULT_CATEGORY_NAME, help="Categoría inteligente a usar.")
    parser.add_argument("--discard-filter", default=DEFAULT_CATEGORY_NAME, help="Filtro/categoría de descarte.")
    parser.add_argument("--temporary-exclusion", default="", help="Término puntual a excluir solo en esta ejecución.")
    parser.add_argument(
        "--scope",
        choices=get_search_scope_options(),
        default=DEFAULT_SEARCH_SCOPE_OPTION,
        help="Alcance de búsqueda.",
    )
    parser.add_argument(
        "--file-type",
        action="append",
        choices=get_search_file_type_options(),
        help="Tipo de archivo a incluir. Puede repetirse.",
    )
    parser.add_argument(
        "--file-types",
        default="",
        help="Tipos de archivo separados por coma, usando los nombres de Smart Filter.",
    )
    parser.add_argument(
        "--all-file-types",
        action="store_true",
        help="Usar todos los tipos de archivo soportados.",
    )
    parser.add_argument("--list-categories", action="store_true", help="Listar categorías disponibles y salir.")
    parser.add_argument("--list-file-types", action="store_true", help="Listar tipos de archivo disponibles y salir.")
    parser.add_argument(
        "--portable-self-check",
        action="store_true",
        help=argparse.SUPPRESS,
    )
    return parser


def _run_portable_self_check() -> int:
    """Exercise bundled engine, grouping, XLSX sanitization, and RenderCore HTML.

    The release build invokes this through SmartFilterCLI.exe. A successful
    check proves that the executable carries the current engine, strict category
    scope, structured XLSX reader, openpyxl, and the Jinja templates required
    by the professional highlighted viewer.
    """

    try:
        from openpyxl import Workbook, load_workbook

        from smart_filter.domain.search_models import FileCandidate, SearchRequest, SearchResult
        from smart_filter.engine.candidate_analysis import CandidateAnalyzer
        from smart_filter.engine.category_scope import extract_category_content_scope
        from smart_filter.readers.xlsx_reader import read_xlsx_text
        from smart_filter.services.document_highlight_service import render_document_highlight
        from smart_filter.services.result_action_service import _create_text_visual_workbook

        for factory_path, config_type in (
            (FACTORY_SETTINGS_PATH, "smartfilter_settings"),
            (FACTORY_CATEGORIES_PATH, "category_database"),
        ):
            if not factory_path.is_file():
                raise RuntimeError(f"Falta default de fábrica empaquetado: {factory_path}")
            factory_document = json.loads(factory_path.read_text(encoding="utf-8-sig"))
            if factory_document.get("meta", {}).get("config_type") != config_type:
                raise RuntimeError(f"Default de fábrica inválido: {factory_path}")
        factory_settings = json.loads(FACTORY_SETTINGS_PATH.read_text(encoding="utf-8-sig"))["data"]
        factory_state = factory_settings.get("state", {})
        if any(
            factory_state.get(key)
            for key in ("last_folder", "last_file", "last_search_text", "last_context_filter")
        ):
            raise RuntimeError("Los defaults empaquetados contienen estado de usuario.")
        if factory_state.get("last_category") != DEFAULT_CATEGORY_NAME:
            raise RuntimeError("Los defaults empaquetados contienen una categoría seleccionada.")

        state = SearchFormState(
            mode=ANALYSIS_MODE_FILE,
            path="portable_self_check.txt",
            search_text="",
            category="portable_self_check",
            search_scope="Nombre y contenido",
            file_types=["Texto (.txt/.log/.md)"],
            source="portable_self_check",
        )
        request = SearchRequest(
            form_state=state,
            category_name="portable_self_check",
            category_terms=["administracion"],
            search_scope="Nombre y contenido",
            extensions=[".txt"],
        )
        candidate = FileCandidate(
            full_path="portable_self_check.txt",
            file_name="portable_self_check.txt",
            extension=".txt",
            folder_path=".",
            content_text="administracion uno\ntexto\nadministracion dos",
            content_reader="text_reader",
            content_status="ok",
            content_chars=45,
        )
        outcome = CandidateAnalyzer(request).analyze(candidate)
        if len(outcome.results) != 1:
            raise RuntimeError(f"Agrupación inesperada: {len(outcome.results)} filas")
        grouped = outcome.results[0]
        if not grouped.grouped_by_file or grouped.occurrence_count != 2:
            raise RuntimeError(
                "El motor incluido no conserva la agrupación por archivo "
                f"(grouped={grouped.grouped_by_file}, occurrences={grouped.occurrence_count})."
            )

        with tempfile.TemporaryDirectory(prefix="smartfilter_portable_self_check_") as temp_dir:
            temp_root = Path(temp_dir)
            source_path = temp_root / "control_chars.log"
            source_text = "normal\nlínea con nulo \x00 y campana \x07 administracion\n"
            source_path.write_bytes(source_text.encode("utf-8"))
            highlight_result = SearchResult(
                index=1,
                candidate=FileCandidate.from_path(
                    source_path,
                    content_text=source_text,
                    source="portable_self_check",
                    content_reader="text_reader",
                    content_status="ok",
                ),
                match_source="Contenido",
                matches="Coincidencia: administracion",
                matched_terms=["administracion"],
            )
            output_path = temp_root / "highlight.xlsx"
            _create_text_visual_workbook(
                source_path,
                highlight_result,
                output_path,
                ["administracion"],
            )
            workbook = load_workbook(output_path, read_only=True, data_only=True)
            try:
                rendered_text = str(workbook.active.cell(row=3, column=2).value or "")
            finally:
                workbook.close()
            if "\x00" in rendered_text or "\x07" in rendered_text:
                raise RuntimeError("La copia destacada conservó caracteres XML no válidos.")

            html_source = temp_root / "highlight_source.xlsx"
            html_workbook = Workbook()
            html_sheet = html_workbook.active
            html_sheet.title = "Datos"
            html_sheet.append(["Área", "Descripción"])
            html_sheet.append(["Administracion", "Liquidación y mantenimiento"])
            html_workbook.save(html_source)
            html_workbook.close()

            structured_text = read_xlsx_text(html_source)
            area_scope = extract_category_content_scope(structured_text, ["Área"])
            if area_scope.text != "Administracion":
                raise RuntimeError(
                    "El reader XLSX o el alcance estructurado no conservaron Área: valor "
                    f"({area_scope.text!r})."
                )
            casual_scope = extract_category_content_scope(
                "Experiencia administrativa requerida para el puesto.",
                ["Experiencia"],
            )
            if casual_scope.found:
                raise RuntimeError("Una frase casual fue interpretada como sección Experiencia.")

            html_result = SearchResult(
                index=1,
                candidate=FileCandidate.from_path(
                    html_source,
                    source="portable_self_check",
                    content_reader="xlsx_reader",
                    content_status="ok",
                ),
                match_source="Contenido",
                matches="Coincidencia: administracion",
                matched_terms=["administracion"],
                category_name="portable_self_check",
                sheet_name="Datos",
                row_number=2,
                location_label="Hoja Datos · Fila 2",
                preview_text="Administracion | Liquidación y mantenimiento",
            )
            html_path = temp_root / "highlight.html"
            html_outcome = render_document_highlight(html_result, html_path)
            if not html_outcome.output_path.is_file():
                raise RuntimeError("RenderCore no generó la vista HTML destacada.")
            html_text = html_outcome.output_path.read_text(encoding="utf-8")
            required_html_markers = (
                "Smart Filter · Documento destacado",
                "administracion",
                "RenderCore",
            )
            missing_markers = [marker for marker in required_html_markers if marker not in html_text]
            if missing_markers:
                raise RuntimeError(
                    "La vista HTML destacada no contiene los marcadores esperados: "
                    + ", ".join(missing_markers)
                )

        print(
            "SMARTFILTER_PORTABLE_SELF_CHECK_OK "
            f"version={APP_VERSION} grouped_occurrences={grouped.occurrence_count} "
            "highlight_xlsx=ok highlight_html=ok section_scope=ok factory_defaults=ok"
        )
        return EXIT_OK
    except Exception as exc:
        print(f"SMARTFILTER_PORTABLE_SELF_CHECK_ERROR version={APP_VERSION} error={exc}")
        return EXIT_ERROR


def _print_lines(lines: list[str], *, quiet: bool) -> None:
    if quiet:
        return
    for line in lines:
        print(line)


def _build_runtime(cli_options) :
    return create_runtime_context(
        tool_name="SmartFilter",
        tool_version=APP_VERSION,
        base_dir=PROJECT_ROOT,
        output_dir=cli_options.output_dir or OUTPUT_DIR,
        logs_dir=cli_options.logs_dir,
        create_directories=True,
    )


def _resolve_output_path(runtime, cli_options) -> Path:
    if cli_options.json_output is None:
        return runtime.get_output_path("smartfilter_results.json")
    candidate = Path(cli_options.json_output).expanduser()
    if candidate.is_absolute():
        return candidate.resolve()
    if cli_options.output_dir is not None:
        return (runtime.output_dir / candidate).resolve()
    return candidate.resolve()


def _selected_file_types(namespace: argparse.Namespace) -> list[str]:
    if namespace.all_file_types:
        return [ALL_FILE_TYPE_OPTION]

    selected: list[str] = []
    valid = set(get_search_file_type_options())
    for item in list(namespace.file_type or []) + _comma_items(namespace.file_types):
        if item in valid and item not in selected:
            selected.append(item)
    return selected or [DEFAULT_FILE_TYPE_OPTION]


def _validate_config_contract(namespace: argparse.Namespace, cli_options) -> int:
    settings_result = load_settings_result()
    categories_result = load_category_result()
    errors = [*settings_result.errors, *categories_result.errors]
    diagnostics = [*settings_result.diagnostics, *categories_result.diagnostics]
    status = "valid" if not errors else "invalid"

    contract = create_result_contract(
        result_type="smartfilter_cli_config_validation",
        tool_name=APP_NAME,
        module_name="SmartFilterCLI",
        extra_meta={"tool_version": APP_VERSION},
        summary={
            "status": status,
            "settings_valid": settings_result.is_valid,
            "categories_valid": categories_result.is_valid,
            "diagnostics_count": len(diagnostics),
            "errors_count": len(errors),
        },
        report_brief={
            "title": "Smart Filter - Validación de configuración CLI",
            "description": "Validación de settings.json y categories.json usando ConfigCore y JsonContractCore.",
        },
        data={
            "settings": settings_result.to_dict(),
            "categories": categories_result.to_dict(),
        },
        diagnostics=[item.to_dict() for item in diagnostics],
        errors=[item.to_dict() for item in errors],
    )

    if cli_options.json_output:
        output_path = Path(cli_options.json_output).expanduser().resolve()
    else:
        output_path = OUTPUT_DIR / "smartfilter_cli_config_validation.json"
    write_json_file(contract, output_path)
    validation = validate_contract(contract, source=str(output_path), strict_schema_version=True)

    _print_lines(
        [
            f"Configuración: {status}",
            f"JSON validación: {output_path}",
            f"Contrato estándar: {'OK' if validation.is_valid else 'ERROR'}",
        ],
        quiet=cli_options.quiet,
    )
    return EXIT_OK if status == "valid" and validation.is_valid else EXIT_ERROR


def _run_search_cli(namespace: argparse.Namespace, cli_options) -> int:
    runtime = _build_runtime(cli_options)
    log_path = runtime.get_log_path(f"smart_filter_{runtime.run_id}.log")
    logger = create_logger("SmartFilterCLI", log_path=log_path, min_level=cli_options.log_level, keep_entries=True)
    logger.info("Inicio de ejecución CLI de Smart Filter.", code="SMARTFILTER_CLI_START")

    category = str(namespace.category or DEFAULT_CATEGORY_NAME).strip() or DEFAULT_CATEGORY_NAME
    if category != DEFAULT_CATEGORY_NAME and category not in get_category_names(include_disabled=True):
        logger.error("La categoría indicada no existe.", code="SMARTFILTER_CATEGORY_NOT_FOUND", context={"category": category})
        _print_lines([f"Categoría no encontrada: {category}"], quiet=cli_options.quiet)
        return EXIT_ERROR

    path: Path | None = namespace.folder or namespace.file
    mode = ANALYSIS_MODE_FOLDER if namespace.folder is not None else ANALYSIS_MODE_FILE
    if path is None:
        logger.error("No se indicó carpeta ni archivo.", code="SMARTFILTER_SOURCE_REQUIRED")
        _print_lines(["Indicar --folder o --file para ejecutar una búsqueda real."], quiet=cli_options.quiet)
        return EXIT_ERROR

    state = SearchFormState(
        mode=mode,
        path=str(path.expanduser()),
        search_text=str(namespace.search_text or "").strip(),
        context_filter=str(getattr(namespace, "context_filter", "") or "").strip(),
        category=category,
        discard_filter=str(namespace.discard_filter or DEFAULT_CATEGORY_NAME).strip() or DEFAULT_CATEGORY_NAME,
        temporary_exclusion=str(namespace.temporary_exclusion or "").strip(),
        search_scope=namespace.scope,
        file_types=_selected_file_types(namespace),
        source="cli",
    )

    if not state.has_search_criteria:
        logger.error("No se indicó criterio de búsqueda.", code="SMARTFILTER_CRITERIA_REQUIRED")
        _print_lines(["Indicar --query y/o --category para buscar."], quiet=cli_options.quiet)
        return EXIT_ERROR

    logger.info(
        "Ejecutando búsqueda CLI.",
        code="SMARTFILTER_SEARCH_RUN",
        context={
            "mode": state.mode,
            "path": state.path,
            "category": state.category,
            "context_filter": state.context_filter,
            "file_types": state.file_types,
            "scope": state.search_scope,
        },
    )
    summary = run_search(state)
    for error in summary.errors:
        logger.warning("La búsqueda finalizó con advertencia/error no fatal.", code="SMARTFILTER_SEARCH_WARNING", context={"error": error})

    output_path = _resolve_output_path(runtime, cli_options)
    write_search_results_contract(
        summary=summary,
        runtime=runtime,
        cli_options=cli_options.to_dict(),
        output_path=output_path,
        logger=logger,
    )
    contract_validation = validate_contract(
        __import__("json").loads(output_path.read_text(encoding="utf-8")),
        source=str(output_path),
        strict_schema_version=True,
    )

    logger.info("Ejecución CLI finalizada.", code="SMARTFILTER_CLI_FINISH", context={"output_path": str(output_path)})
    _print_lines(
        [
            f"Resultados encontrados: {summary.match_occurrences_count}",
            f"Archivos coincidentes: {summary.matched_candidates_count}",
            f"JSON estándar: {output_path}",
            f"Log: {log_path}",
            f"Contrato estándar: {'OK' if contract_validation.is_valid else 'ERROR'}",
        ],
        quiet=cli_options.quiet,
    )
    return EXIT_OK if contract_validation.is_valid else EXIT_ERROR


def run_cli(argv: Sequence[str] | None = None) -> int:
    parser = build_parser()
    effective_argv = list(argv) if argv is not None else sys.argv[1:]
    if not effective_argv:
        parser.print_help()
        return EXIT_OK

    ensure_project_directories()
    namespace = parser.parse_args(effective_argv)
    cli_options = parse_cli_options(parser, effective_argv)

    if namespace.list_categories:
        _print_lines(get_category_names(include_disabled=True), quiet=cli_options.quiet)
        return EXIT_OK

    if namespace.list_file_types:
        _print_lines(get_search_file_type_options(), quiet=cli_options.quiet)
        return EXIT_OK

    if namespace.validate_config:
        return _validate_config_contract(namespace, cli_options)

    if namespace.portable_self_check:
        return _run_portable_self_check()

    return _run_search_cli(namespace, cli_options)
