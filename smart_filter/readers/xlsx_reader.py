from __future__ import annotations

from pathlib import Path
from typing import Any


def _clean_row(row: tuple[Any, ...]) -> list[str]:
    return [str(value).strip() if value is not None else "" for value in row]


def _looks_like_header(row: list[str], following_rows: list[list[str]]) -> bool:
    nonempty = [value for value in row if value]
    if len(nonempty) < 2 or not following_rows:
        return False
    if len(set(value.casefold() for value in nonempty)) != len(nonempty):
        return False
    if not all(len(value) <= 60 and any(character.isalpha() for character in value) for value in nonempty):
        return False

    # A header is more credible when the following row differs in shape/type or
    # contains longer/free-form values. This stays conservative for generic sheets.
    second = [value for value in following_rows[0] if value]
    return bool(second)


def _worksheet_lines(worksheet) -> list[str]:
    rows = [_clean_row(row) for row in worksheet.iter_rows(values_only=True)]
    rows = [row for row in rows if any(row)]
    if not rows:
        return []

    output: list[str] = []
    if worksheet.title:
        output.append(f"Hoja: {worksheet.title}")

    headers = rows[0]
    if _looks_like_header(headers, rows[1:]):
        for row in rows[1:]:
            emitted = False
            for index, value in enumerate(row):
                if not value:
                    continue
                header = headers[index].strip() if index < len(headers) else ""
                output.append(f"{header}: {value}" if header else value)
                emitted = True
            if emitted:
                output.append("")
    else:
        for row in rows:
            values = [value for value in row if value]
            if values:
                output.append(" | ".join(values))

    return output


def read_xlsx_text(file_path: str | Path) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - depends on local environment.
        raise RuntimeError("Falta dependencia openpyxl para leer archivos .xlsx") from exc

    workbook = None
    try:
        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        extracted_text: list[str] = []
        for worksheet in workbook.worksheets:
            extracted_text.extend(_worksheet_lines(worksheet))
        return "\n".join(extracted_text).strip()
    finally:
        if workbook is not None:
            workbook.close()
