from __future__ import annotations

import csv
import html
import json
import os
import shutil
import subprocess
import sys
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable, Mapping

from date_time_core import create_timestamp_pair, datetime_to_run_id_timestamp, utc_now
from platform_core import open_path as platform_open_path

from smart_filter.domain.search_models import (
    FileCandidate,
    SearchRequest,
    SearchResult,
    SearchSummary,
    count_unique_result_candidates,
)
from smart_filter.domain.text_normalizer import normalize_text
from smart_filter.domain.search_form_state import SearchFormState
from smart_filter.paths import OUTPUT_DIR, TEMP_DIR, ensure_project_directories
from smart_filter.services.settings_service import (
    get_highlight_cell_color_palette,
    get_highlight_text_color_hex,
    get_settings,
)
from smart_filter.services.document_highlight_service import render_document_highlight

OUTPUT_MARKER_FILE_NAME = ".smartfilter_output"
DEFAULT_EXPORT_PREFIX = "SmartFilter_Resultados"
DEFAULT_HIGHLIGHT_FOLDER = "highlight_previews"
DEFAULT_HIGHLIGHT_COPY_FOLDER = "highlight_copies"
TEXT_LIKE_EXTENSIONS = {".txt", ".log", ".md", ".csv", ".json", ".xml", ".html", ".htm"}
EXCEL_ILLEGAL_XML_CHARACTERS_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")
OFFICE_ORIGINAL_EXTENSIONS = {".doc", ".docx", ".xls", ".xlsm", ".xlsx"}


@dataclass(frozen=True)
class ResultActionOutcome:
    success: bool
    message: str
    path: str = ""
    details: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "success": self.success,
            "message": self.message,
            "path": self.path,
            "details": list(self.details),
        }


@dataclass(frozen=True)
class ExportOutcome:
    success: bool
    output_folder: str
    csv_path: str = ""
    json_path: str = ""
    exported_count: int = 0
    errors: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "success": self.success,
            "output_folder": self.output_folder,
            "csv_path": self.csv_path,
            "json_path": self.json_path,
            "exported_count": self.exported_count,
            "errors": list(self.errors),
        }


@dataclass(frozen=True)
class PreviousResultsViewOutcome:
    success: bool
    message: str
    view_folder: str
    prefix: str
    matched_folders_count: int = 0
    result_files_count: int = 0
    errors: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "success": self.success,
            "message": self.message,
            "view_folder": self.view_folder,
            "prefix": self.prefix,
            "matched_folders_count": self.matched_folders_count,
            "result_files_count": self.result_files_count,
            "errors": list(self.errors),
        }


@dataclass(frozen=True)
class ImportResultsOutcome:
    success: bool
    message: str
    source_path: str
    summary: SearchSummary | None = None
    imported_count: int = 0
    errors: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "success": self.success,
            "message": self.message,
            "source_path": self.source_path,
            "imported_count": self.imported_count,
            "errors": list(self.errors),
            "summary": self.summary.to_dict() if self.summary is not None else None,
        }


def _timestamp() -> str:
    return datetime_to_run_id_timestamp(utc_now())


def _safe_name(value: str, fallback: str = "archivo") -> str:
    clean = re.sub(r"[^A-Za-z0-9._-]+", "_", str(value or "").strip())
    return clean.strip("._-") or fallback


def _settings_output_prefix() -> str:
    settings = get_settings()
    prefix = str(settings.get("output_folder_prefix") or DEFAULT_EXPORT_PREFIX).strip()
    return _safe_name(prefix, DEFAULT_EXPORT_PREFIX)


def ensure_output_marker(folder: str | Path) -> None:
    marker_path = Path(folder) / OUTPUT_MARKER_FILE_NAME
    try:
        marker_path.write_text(
            "Smart Filter output folder. This folder can be skipped during parent scans.\n",
            encoding="utf-8",
        )
    except Exception:
        pass


def create_action_output_folder(prefix: str | None = None) -> Path:
    ensure_project_directories()
    clean_prefix = _safe_name(prefix or _settings_output_prefix(), DEFAULT_EXPORT_PREFIX)
    base = OUTPUT_DIR / f"{clean_prefix}_{_timestamp()}"
    if not base.exists():
        base.mkdir(parents=True, exist_ok=False)
        ensure_output_marker(base)
        return base

    counter = 1
    while True:
        candidate = OUTPUT_DIR / f"{clean_prefix}_{_timestamp()}_{counter:03d}"
        if not candidate.exists():
            candidate.mkdir(parents=True, exist_ok=False)
            ensure_output_marker(candidate)
            return candidate
        counter += 1


def _is_result_file(path: Path) -> bool:
    if not path.is_file():
        return False
    return "resultado" in path.stem.lower()


def _iter_previous_result_folders(prefix: str) -> list[Path]:
    if not OUTPUT_DIR.exists():
        return []
    expected_prefix = f"{prefix}_"
    folders: list[Path] = []
    try:
        for child in OUTPUT_DIR.iterdir():
            if not child.is_dir():
                continue
            if not child.name.startswith(expected_prefix):
                continue
            folders.append(child)
    except Exception:
        return []
    return sorted(folders, key=lambda item: item.stat().st_mtime if item.exists() else 0, reverse=True)


def build_previous_results_view() -> PreviousResultsViewOutcome:
    """Build a clean folder view for previous result exports.

    The user-facing action should not open the raw output directory because it can
    contain snapshots, markers, logs or auxiliary files. This view only mirrors
    folders that match the configured output prefix and only copies files whose
    name is result-focused, such as resultados.csv or resultados.json.
    """

    ensure_project_directories()
    prefix = _settings_output_prefix()
    view_root = TEMP_DIR / "previous_results" / prefix
    errors: list[str] = []

    try:
        if view_root.exists():
            shutil.rmtree(view_root, ignore_errors=True)
        view_root.mkdir(parents=True, exist_ok=True)
    except Exception as exc:
        return PreviousResultsViewOutcome(
            False,
            f"No fue posible preparar la vista filtrada: {exc}",
            str(view_root),
            prefix,
            errors=[str(exc)],
        )

    matched_folders = 0
    copied_files = 0
    for source_folder in _iter_previous_result_folders(prefix):
        result_files = [path for path in sorted(source_folder.iterdir()) if _is_result_file(path)]
        if not result_files:
            continue

        target_folder = view_root / source_folder.name
        try:
            target_folder.mkdir(parents=True, exist_ok=True)
        except Exception as exc:
            errors.append(f"{source_folder.name}: {exc}")
            continue

        folder_copied = 0
        for result_file in result_files:
            target_file = target_folder / result_file.name
            try:
                shutil.copy2(result_file, target_file)
                folder_copied += 1
                copied_files += 1
            except Exception as exc:
                errors.append(f"{result_file}: {exc}")
        if folder_copied:
            matched_folders += 1

    if not copied_files:
        return PreviousResultsViewOutcome(
            False,
            "No se encontraron carpetas de resultados con el prefijo configurado ni archivos de resultados para mostrar.",
            str(view_root),
            prefix,
            matched_folders,
            copied_files,
            errors,
        )

    return PreviousResultsViewOutcome(
        True,
        "Vista filtrada de resultados previos preparada.",
        str(view_root),
        prefix,
        matched_folders,
        copied_files,
        errors,
    )


def _coerce_int(value: Any) -> int | None:
    try:
        if value is None or str(value).strip() == "":
            return None
        return int(float(str(value).strip()))
    except Exception:
        return None


def _row_value(row: Mapping[str, Any], *names: str, default: str = "") -> str:
    for name in names:
        if name in row and row.get(name) is not None:
            value = str(row.get(name) or "").strip()
            if value:
                return value
    return default


def _split_exported_terms(value: Any) -> list[str]:
    if isinstance(value, list):
        raw_terms = value
    else:
        raw_text = str(value or "")
        raw_terms = re.split(r"\s*\|\s*|\s*,\s*|\s*;\s*", raw_text)
    terms: list[str] = []
    seen: set[str] = set()
    for term in raw_terms:
        clean = str(term or "").strip()
        key = normalize_text(clean)
        if not clean or not key or key in seen:
            continue
        seen.add(key)
        terms.append(clean)
    return terms


def _read_import_rows(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    errors: list[str] = []
    suffix = path.suffix.lower()
    if suffix == ".json":
        try:
            data = json.loads(path.read_text(encoding="utf-8-sig"))
        except Exception as exc:
            return [], [f"JSON: {exc}"]
        if isinstance(data, dict):
            rows = data.get("exported_results")
            if rows is None and isinstance(data.get("summary"), dict):
                rows = data.get("summary", {}).get("results")
            if rows is None:
                rows = data.get("results")
        elif isinstance(data, list):
            rows = data
        else:
            rows = []
        if not isinstance(rows, list):
            return [], ["El archivo JSON no contiene una lista de resultados reconocible."]
        return [dict(item) for item in rows if isinstance(item, Mapping)], errors

    if suffix == ".csv":
        try:
            raw_text = path.read_text(encoding="utf-8-sig")
        except Exception:
            try:
                raw_text = path.read_text(encoding="utf-8", errors="replace")
            except Exception as exc:
                return [], [f"CSV: {exc}"]
        try:
            sample = raw_text[:4096]
            dialect = csv.Sniffer().sniff(sample, delimiters=";,\t,") if sample.strip() else csv.excel
        except Exception:
            class _SmartFilterCsvDialect(csv.excel):
                delimiter = ";"
            dialect = _SmartFilterCsvDialect
        try:
            reader = csv.DictReader(raw_text.splitlines(), dialect=dialect)
            return [dict(row) for row in reader if row], errors
        except Exception as exc:
            return [], [f"CSV: {exc}"]

    return [], ["Formato no compatible. Usar resultados.json o resultados.csv."]


def _parse_imported_locations(value: Any) -> list[dict[str, Any]]:
    if isinstance(value, list):
        return [dict(item) for item in value if isinstance(item, Mapping)]
    text = str(value or "").strip()
    if not text:
        return []
    try:
        parsed = json.loads(text)
    except Exception:
        return []
    if not isinstance(parsed, list):
        return []
    return [dict(item) for item in parsed if isinstance(item, Mapping)]


def _import_row_to_result(row: Mapping[str, Any], fallback_index: int, used_indexes: set[int]) -> SearchResult | None:
    raw_path = _row_value(row, "ruta", "full_path", "path")
    file_name = _row_value(row, "archivo", "file_name", default=Path(raw_path).name if raw_path else f"resultado_{fallback_index:03d}")
    folder_path = _row_value(row, "carpeta", "folder_path", default=str(Path(raw_path).parent) if raw_path else "")
    extension = _row_value(row, "extension", "file_type", default=Path(file_name).suffix.lower())
    if extension == "sin extensión":
        extension = ""
    full_path = raw_path or (str(Path(folder_path) / file_name) if folder_path else file_name)

    index = _coerce_int(_row_value(row, "numero", "index")) or fallback_index
    while index in used_indexes:
        index += 1
    used_indexes.add(index)

    content_chars = _coerce_int(_row_value(row, "caracteres_contenido", "content_chars")) or 0
    candidate = FileCandidate(
        full_path=full_path,
        file_name=file_name,
        extension=extension.lower(),
        folder_path=folder_path,
        content_text="",
        source="imported_results",
        size_bytes=Path(full_path).stat().st_size if full_path and Path(full_path).exists() and Path(full_path).is_file() else None,
        content_reader=_row_value(row, "reader", "content_reader", default="imported"),
        content_status=_row_value(row, "estado_contenido", "content_status", default="importado"),
        content_chars=content_chars,
    )
    return SearchResult(
        index=index,
        candidate=candidate,
        match_source=_row_value(row, "coincidio_en", "match_source", "source", default="Importado"),
        matches=_row_value(row, "coincidencias", "matches", "match"),
        matched_terms=_split_exported_terms(row.get("terminos_encontrados", row.get("matched_terms", row.get("terms", "")))),
        status=_row_value(row, "estado", "status", default="Importado"),
        category_name=_row_value(row, "categoria", "category_name", "category", default="Ninguna"),
        discard_filter_name=_row_value(row, "filtro_descarte", "discard_filter_name", "discard", default="Ninguna"),
        occurrence_number=_coerce_int(_row_value(row, "occurrence_number", "ocurrencia")),
        line_number=_coerce_int(_row_value(row, "linea", "line_number")),
        row_number=_coerce_int(_row_value(row, "row_number", "fila")),
        sheet_name=_row_value(row, "sheet_name", "hoja"),
        location_label=_row_value(row, "ubicacion", "location_label", "location", default="Archivo"),
        preview_text=_row_value(row, "vista_previa", "preview_text", "preview"),
        occurrence_count=_coerce_int(_row_value(row, "ocurrencias", "occurrence_count")) or 1,
        match_locations=_parse_imported_locations(row.get("ubicaciones_detalle", row.get("match_locations", []))),
        grouped_by_file=str(row.get("agrupado_por_archivo", row.get("grouped_by_file", ""))).strip().lower() in {"1", "true", "si", "sí", "yes"},
    )


def import_results_file(path: str | Path) -> ImportResultsOutcome:
    source = Path(path)
    if not source.exists() or not source.is_file():
        return ImportResultsOutcome(False, "El archivo de resultados no existe.", str(source))

    rows, errors = _read_import_rows(source)
    if errors:
        return ImportResultsOutcome(False, "No fue posible leer el archivo de resultados.", str(source), errors=errors)
    if not rows:
        return ImportResultsOutcome(False, "El archivo no contiene resultados para importar.", str(source))

    imported: list[SearchResult] = []
    used_indexes: set[int] = set()
    for fallback_index, row in enumerate(rows, start=1):
        try:
            result = _import_row_to_result(row, fallback_index, used_indexes)
            if result is not None:
                imported.append(result)
        except Exception as exc:
            errors.append(f"Fila {fallback_index}: {exc}")

    if not imported:
        return ImportResultsOutcome(False, "No se pudo reconstruir ningún resultado importado.", str(source), errors=errors)

    first = imported[0]
    form_state = SearchFormState(
        mode="Archivo",
        path=str(source),
        search_text="Resultados importados",
        category=first.category_name or "Ninguna",
        discard_filter=first.discard_filter_name or "Ninguna",
        search_scope="Nombre y contenido",
        file_types=["Todos los compatibles"],
        remember_last_location=True,
        save_search_history=True,
        remember_last_search_settings=True,
        source="imported_results",
    )
    request = SearchRequest(
        form_state=form_state,
        search_text="Resultados importados",
        category_name=form_state.category,
        discard_filter_name=form_state.discard_filter,
        search_scope=form_state.search_scope,
        file_types=list(form_state.file_types),
    )
    matched_candidates_count = count_unique_result_candidates(imported)
    summary = SearchSummary(
        request=request,
        results=imported,
        analyzed_candidates_count=matched_candidates_count,
        matched_candidates_count=matched_candidates_count,
        no_match_count=0,
        skipped_by_discard_count=0,
        unsupported_extension_count=0,
        errors=errors,
        scan_stats={
            "source": "imported_results",
            "import_path": str(source),
            "candidates_count": matched_candidates_count,
            "match_occurrences_count": sum(max(1, int(item.occurrence_count or 1)) for item in imported),
            "readers_executed_count": 0,
            "content_text_chars_count": 0,
        },
    )
    return ImportResultsOutcome(
        success=True,
        message="Resultados importados correctamente.",
        source_path=str(source),
        summary=summary,
        imported_count=len(imported),
        errors=errors,
    )


def open_path(path: str | Path) -> ResultActionOutcome:
    target = Path(path)
    if not target.exists():
        return ResultActionOutcome(False, "La ruta no existe.", str(target))

    try:
        platform_open_path(target)
        return ResultActionOutcome(True, "Ruta abierta correctamente.", str(target))
    except Exception as exc:
        return ResultActionOutcome(
            False,
            f"No se pudo abrir el archivo con la aplicación predeterminada: {exc}",
            str(target),
        )


def _open_original_office_path(target: Path) -> None:
    """Open the exact Office document path without creating a working copy.

    PlatformCore's generic opener remains appropriate for folders and regular
    files. Office documents use the operating system association directly so
    Excel and Word receive the original path selected in the results table.
    """

    if sys.platform == "win32":
        startfile = getattr(os, "startfile", None)
        if startfile is None:
            raise RuntimeError("Windows no ofrece os.startfile en este entorno.")
        startfile(str(target))
        return

    if sys.platform.startswith("linux"):
        subprocess.Popen(
            ["xdg-open", str(target)],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            close_fds=True,
        )
        return

    platform_open_path(target)


def open_original(result: SearchResult) -> ResultActionOutcome:
    target = Path(result.full_path)
    if not target.exists() or not target.is_file():
        return ResultActionOutcome(False, "El archivo original no existe.", str(target))

    try:
        if target.suffix.lower() in OFFICE_ORIGINAL_EXTENSIONS:
            _open_original_office_path(target)
        else:
            platform_open_path(target)
        return ResultActionOutcome(True, "Archivo original abierto correctamente.", str(target))
    except Exception as exc:
        return ResultActionOutcome(
            False,
            f"No se pudo abrir el archivo original con la aplicación predeterminada: {exc}",
            str(target),
        )


def open_parent_folder(result: SearchResult) -> ResultActionOutcome:
    return open_path(result.folder_path)


def copy_original_files(results: Iterable[SearchResult], *, preserve_folder_structure: bool = False) -> ExportOutcome:
    selected_results = list(results)
    output_folder = create_action_output_folder("SmartFilter_Copias")
    errors: list[str] = []
    copied = 0

    for result in selected_results:
        source = Path(result.full_path)
        if not source.exists() or not source.is_file():
            errors.append(f"No existe: {source}")
            continue

        destination_folder = output_folder
        if preserve_folder_structure:
            try:
                root = Path(result.folder_path).resolve()
                relative = source.resolve().parent.relative_to(root)
                destination_folder = output_folder / relative
                destination_folder.mkdir(parents=True, exist_ok=True)
            except Exception:
                destination_folder = output_folder

        destination = destination_folder / source.name
        if destination.exists():
            destination = destination_folder / f"{source.stem}_{copied + 1:03d}{source.suffix}"

        try:
            shutil.copy2(source, destination)
            copied += 1
        except Exception as exc:
            errors.append(f"{source}: {exc}")

    return ExportOutcome(
        success=copied > 0 and not errors,
        output_folder=str(output_folder),
        exported_count=copied,
        errors=errors,
    )


def build_paths_clipboard_text(results: Iterable[SearchResult]) -> str:
    return "\n".join(result.full_path for result in results)


def build_results_clipboard_text(results: Iterable[SearchResult]) -> str:
    lines = ["#\tEstado\tArchivo\tTipo\tCoincidencia\tTérminos\tRuta"]
    for result in results:
        lines.append(
            "\t".join(
                [
                    str(result.index),
                    result.status,
                    result.file_name,
                    result.extension or "sin extensión",
                    result.location_label or "Archivo",
                    result.matches,
                    " | ".join(result.matched_terms),
                    result.preview_text,
                    result.full_path,
                ]
            )
        )
    return "\n".join(lines)


def _result_to_export_row(result: SearchResult) -> dict[str, Any]:
    return {
        "numero": result.index,
        "estado": result.status,
        "archivo": result.file_name,
        "extension": result.extension,
        "categoria": result.category_name,
        "filtro_descarte": result.discard_filter_name,
        "coincidio_en": result.match_source,
        "ubicacion": result.location_label or "Archivo",
        "linea": result.line_number or "",
        "coincidencias": result.matches,
        "terminos_encontrados": " | ".join(result.matched_terms),
        "vista_previa": result.preview_text,
        "ocurrencias": max(1, int(result.occurrence_count or 1)),
        "agrupado_por_archivo": bool(result.grouped_by_file),
        "ubicaciones_detalle": json.dumps(result.match_locations, ensure_ascii=False),
        "reader": result.candidate.content_reader,
        "estado_contenido": result.candidate.content_status,
        "caracteres_contenido": result.candidate.content_chars,
        "carpeta": result.folder_path,
        "ruta": result.full_path,
    }


def _format_elapsed_seconds(value: Any) -> str:
    try:
        total_seconds = max(0.0, float(value))
    except (TypeError, ValueError):
        total_seconds = 0.0
    hours, remainder = divmod(total_seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    if hours >= 1:
        return f"{int(hours):02d}:{int(minutes):02d}:{seconds:06.3f}"
    return f"{int(minutes):02d}:{seconds:06.3f}"


def _search_timing_payload(summary: SearchSummary) -> dict[str, Any]:
    stats = dict(summary.scan_stats or {})
    integrated_seconds = float(
        stats.get("scan_read_match_elapsed_seconds")
        or stats.get("scan_and_read_elapsed_seconds")
        or 0.0
    )
    scan_seconds = float(stats.get("scan_and_read_elapsed_seconds") or integrated_seconds)
    analysis_seconds = float(stats.get("match_analysis_elapsed_seconds") or 0.0)
    worker_analysis_seconds = float(stats.get("match_analysis_worker_elapsed_seconds_total") or 0.0)
    total_seconds = float(stats.get("total_search_elapsed_seconds") or 0.0)
    return {
        "execution_pipeline_mode": stats.get("execution_pipeline_mode", "legacy_separate_match"),
        "match_analysis_integrated": bool(stats.get("match_analysis_integrated", False)),
        "scan_read_match_elapsed_seconds": round(integrated_seconds, 6),
        "scan_read_match_duration": _format_elapsed_seconds(integrated_seconds),
        "scan_and_read_elapsed_seconds": round(scan_seconds, 6),
        "scan_and_read_duration": _format_elapsed_seconds(scan_seconds),
        "scan_and_read_includes_match_analysis": bool(
            stats.get("scan_and_read_includes_match_analysis", False)
        ),
        "match_analysis_elapsed_seconds": round(analysis_seconds, 6),
        "match_analysis_duration": _format_elapsed_seconds(analysis_seconds),
        "match_analysis_worker_elapsed_seconds_total": round(worker_analysis_seconds, 6),
        "match_analysis_separate_pass_eliminated": bool(
            stats.get("match_analysis_separate_pass_eliminated", False)
        ),
        "total_search_elapsed_seconds": round(total_seconds, 6),
        "total_search_duration": _format_elapsed_seconds(total_seconds),
    }


def export_results(summary: SearchSummary, *, selected_results: Iterable[SearchResult] | None = None) -> ExportOutcome:
    results = list(selected_results) if selected_results is not None else list(summary.results)
    output_folder = create_action_output_folder()
    csv_path = output_folder / "resultados.csv"
    json_path = output_folder / "resultados.json"
    errors: list[str] = []

    try:
        fields = [
            "numero",
            "estado",
            "archivo",
            "extension",
            "categoria",
            "filtro_descarte",
            "coincidio_en",
            "ubicacion",
            "linea",
            "coincidencias",
            "terminos_encontrados",
            "vista_previa",
            "ocurrencias",
            "agrupado_por_archivo",
            "ubicaciones_detalle",
            "reader",
            "estado_contenido",
            "caracteres_contenido",
            "carpeta",
            "ruta",
        ]
        with csv_path.open("w", newline="", encoding="utf-8-sig") as file:
            writer = csv.DictWriter(file, fieldnames=fields, delimiter=";")
            writer.writeheader()
            for result in results:
                writer.writerow(_result_to_export_row(result))
    except Exception as exc:
        errors.append(f"CSV: {exc}")

    try:
        exported_at_utc, exported_at_local = create_timestamp_pair()
        payload = {
            "export": {
                "generated_at_local": exported_at_local,
                "generated_at_utc": exported_at_utc,
            },
            "timing": _search_timing_payload(summary),
            "performance": dict(summary.scan_stats.get("performance") or {}),
            "summary": summary.to_dict(),
            "exported_results": [_result_to_export_row(result) for result in results],
        }
        json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception as exc:
        errors.append(f"JSON: {exc}")

    return ExportOutcome(
        success=not errors,
        output_folder=str(output_folder),
        csv_path=str(csv_path),
        json_path=str(json_path),
        exported_count=len(results),
        errors=errors,
    )


def _highlight_text_fragment(text: str, terms: Iterable[str]) -> str:
    escaped = html.escape(text or "")
    unique_terms: list[str] = []
    seen: set[str] = set()
    for term in terms:
        clean = str(term or "").strip()
        key = normalize_text(clean)
        if not clean or not key or key in seen:
            continue
        seen.add(key)
        unique_terms.append(clean)

    for term in sorted(unique_terms, key=len, reverse=True):
        pattern = re.compile(re.escape(html.escape(term)), re.IGNORECASE)
        escaped = pattern.sub(lambda match: f"<mark>{match.group(0)}</mark>", escaped)
    return escaped




def sanitize_excel_cell_text(value: Any) -> str:
    """Return text that can be written safely to an XLSX cell.

    Text readers may preserve control characters found in logs, antivirus
    reports, copied console output, or partially binary files. XML 1.0, which
    backs XLSX worksheets, rejects those characters. Replacing them here keeps
    the highlighted copy readable without modifying the original document.
    """

    text = str(value or "")
    if not text:
        return ""
    return EXCEL_ILLEGAL_XML_CHARACTERS_RE.sub("�", text)[:32767]


def _term_matches_text(text: Any, terms: Iterable[str]) -> bool:
    normalized_text = normalize_text(str(text or ""))
    if not normalized_text:
        return False
    for term in terms:
        clean = str(term or "").strip()
        if clean and normalize_text(clean) in normalized_text:
            return True
    return False


def _clean_highlight_terms(result: SearchResult) -> list[str]:
    terms: list[str] = []
    seen: set[str] = set()
    for term in result.matched_terms:
        clean = str(term or "").strip()
        lookup = normalize_text(clean)
        if not clean or not lookup or lookup in seen:
            continue
        seen.add(lookup)
        terms.append(clean)
    return terms


def _highlight_style() -> tuple[Any, Any, Any]:
    try:
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:  # pragma: no cover - depends on local environment.
        raise RuntimeError("Falta dependencia openpyxl para crear copias destacadas") from exc

    settings = get_settings()
    palette = get_highlight_cell_color_palette(str(settings.get("highlight_cell_color") or "Amarillo"))
    text_color = get_highlight_text_color_hex(str(settings.get("highlight_text_color") or "Negro"))
    row_fill_hex = palette.get("row_fill") or "FFF2CC"
    cell_fill_hex = palette.get("cell_fill") or "FFD966"
    row_fill = PatternFill(fill_type="solid", fgColor=row_fill_hex)
    cell_fill = PatternFill(fill_type="solid", fgColor=cell_fill_hex)
    font = Font(color=text_color, bold=True) if text_color else Font(bold=True)
    return row_fill, cell_fill, font


def _autosize_worksheet(worksheet: Any, *, max_width: int = 90) -> None:
    for column_cells in worksheet.columns:
        try:
            column_letter = column_cells[0].column_letter
        except Exception:
            continue
        values = [str(cell.value or "") for cell in column_cells[:120]]
        width = min(max_width, max([len(value) for value in values] + [10]) + 2)
        worksheet.column_dimensions[column_letter].width = width


def _create_text_visual_workbook(source: Path, result: SearchResult, destination: Path, terms: list[str]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:  # pragma: no cover - depends on local environment.
        raise RuntimeError("Falta dependencia openpyxl para crear copias destacadas") from exc

    row_fill, cell_fill, highlight_font = _highlight_style()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Vista destacada"
    worksheet.append(["Línea", "Contenido"])
    header_fill = PatternFill(fill_type="solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    content_text = result.candidate.content_text or ""
    if not content_text and source.exists():
        try:
            content_text = source.read_text(encoding="utf-8", errors="replace")
        except Exception:
            content_text = result.preview_text or result.matches or result.file_name
    lines = content_text.splitlines() or [content_text]

    for line_number, line_text in enumerate(lines, start=1):
        safe_line_text = sanitize_excel_cell_text(line_text)
        worksheet.append([line_number, safe_line_text])
        current_row = worksheet.max_row
        line_matches = _term_matches_text(safe_line_text, terms)
        selected_line = result.line_number is not None and line_number == result.line_number
        if line_matches or selected_line:
            for cell in worksheet[current_row]:
                cell.fill = row_fill
            worksheet.cell(current_row, 2).fill = cell_fill
            worksheet.cell(current_row, 2).font = highlight_font

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    _autosize_worksheet(worksheet)
    workbook.save(destination)


def _create_xlsx_highlight_copy(source: Path, destination: Path, terms: list[str]) -> int:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - depends on local environment.
        raise RuntimeError("Falta dependencia openpyxl para crear copias destacadas") from exc

    shutil.copy2(source, destination)
    row_fill, cell_fill, highlight_font = _highlight_style()
    workbook = load_workbook(destination)
    highlighted = 0
    try:
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                row_matches = False
                matching_cells = []
                for cell in row:
                    if _term_matches_text(cell.value, terms):
                        row_matches = True
                        matching_cells.append(cell)
                if row_matches:
                    highlighted += 1
                    for cell in row:
                        cell.fill = row_fill
                    for cell in matching_cells:
                        cell.fill = cell_fill
                        cell.font = highlight_font
        workbook.save(destination)
    finally:
        try:
            workbook.close()
        except Exception:
            pass
    return highlighted


def create_highlighted_file_copy(result: SearchResult) -> ResultActionOutcome:
    """Create and open a temporary visual copy with matching terms highlighted.

    HTML preview remains available from the dedicated Destacado action. This copy
    is for the normal Abrir action when settings request a highlighted opening,
    matching the older product behavior where a selected file opened as a visual
    copy without modifying the original file.
    """

    source = Path(result.full_path)
    if not source.exists():
        return ResultActionOutcome(False, "El archivo original ya no existe.", str(source))

    terms = _clean_highlight_terms(result)
    if not terms:
        terms = [result.matches] if result.matches else []
    if not terms:
        return ResultActionOutcome(False, "No hay términos disponibles para destacar.", str(source))

    copy_folder = TEMP_DIR / DEFAULT_HIGHLIGHT_COPY_FOLDER
    copy_folder.mkdir(parents=True, exist_ok=True)

    extension = source.suffix.lower()
    safe_stem = _safe_name(source.stem)
    if extension == ".xlsx":
        destination = copy_folder / f"{safe_stem}_copia_destacada_{_timestamp()}.xlsx"
        try:
            highlighted = _create_xlsx_highlight_copy(source, destination, terms)
        except Exception as exc:
            destination.unlink(missing_ok=True)
            return ResultActionOutcome(False, f"No se pudo crear la copia destacada: {exc}", str(destination))
        if highlighted == 0:
            return ResultActionOutcome(False, "No se encontraron celdas para destacar en la copia.", str(destination))
    else:
        destination = copy_folder / f"{safe_stem}_vista_destacada_{_timestamp()}.xlsx"
        try:
            _create_text_visual_workbook(source, result, destination, terms)
        except Exception as exc:
            destination.unlink(missing_ok=True)
            return ResultActionOutcome(False, f"No se pudo crear la vista destacada en Excel: {exc}", str(destination))

    open_result = open_path(destination)
    if not open_result.success:
        return ResultActionOutcome(False, open_result.message, str(destination))
    return ResultActionOutcome(True, "Copia destacada generada y abierta.", str(destination))

def create_highlight_preview(result: SearchResult) -> ResultActionOutcome:
    """Generate the full semantic HTML viewer through SharedCode RenderCore."""

    source = Path(result.full_path)
    if not source.exists():
        return ResultActionOutcome(False, "El archivo original ya no existe.", str(source))

    preview_folder = TEMP_DIR / DEFAULT_HIGHLIGHT_FOLDER
    preview_folder.mkdir(parents=True, exist_ok=True)
    preview_path = preview_folder / f"{_safe_name(source.stem)}_destacado_{_timestamp()}.html"

    try:
        render_outcome = render_document_highlight(result, preview_path)
    except Exception as exc:
        return ResultActionOutcome(
            False,
            f"No se pudo generar la vista destacada con RenderCore: {exc}",
            str(preview_path),
        )

    open_result = open_path(render_outcome.output_path)
    if not open_result.success:
        return ResultActionOutcome(False, open_result.message, str(render_outcome.output_path))

    message = "Documento HTML destacado generado y abierto."
    if render_outcome.truncated:
        message += " La vista fue limitada por seguridad; el original permanece completo."
    return ResultActionOutcome(
        True,
        message,
        str(render_outcome.output_path),
        details=list(render_outcome.diagnostics),
    )


def row_index_value(row: Mapping[str, Any] | None) -> int | None:
    if not row:
        return None
    value = row.get("index")
    try:
        return int(str(value).strip())
    except Exception:
        return None
