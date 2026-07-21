from __future__ import annotations

import tempfile
from pathlib import Path

from smart_filter.bootstrap import ensure_sharedcode_on_path

ensure_sharedcode_on_path()

from openpyxl import load_workbook

from smart_filter.domain.search_models import FileCandidate, SearchResult
from smart_filter.services.result_action_service import (
    _create_text_visual_workbook,
    sanitize_excel_cell_text,
)


def main() -> int:
    raw = "línea con nulo \x00, campana \x07 y término administracion"
    cleaned = sanitize_excel_cell_text(raw)
    assert "\x00" not in cleaned
    assert "\x07" not in cleaned
    assert "administracion" in cleaned

    with tempfile.TemporaryDirectory(prefix="smartfilter_highlight_control_chars_") as temp_dir:
        temp_root = Path(temp_dir)
        source = temp_root / "antivirus.log"
        content = "cabecera\n" + raw + "\nfin"
        source.write_bytes(content.encode("utf-8"))
        result = SearchResult(
            index=1,
            candidate=FileCandidate.from_path(
                source,
                content_text=content,
                source="validator",
                content_reader="text_reader",
                content_status="ok",
            ),
            match_source="Contenido",
            matches="Coincidencia: administracion",
            matched_terms=["administracion"],
            line_number=2,
        )
        output = temp_root / "vista_destacada.xlsx"
        _create_text_visual_workbook(source, result, output, ["administracion"])
        assert output.is_file() and output.stat().st_size > 0

        workbook = load_workbook(output, read_only=True, data_only=True)
        try:
            cell_text = str(workbook.active.cell(row=3, column=2).value or "")
        finally:
            workbook.close()
        assert "\x00" not in cell_text
        assert "\x07" not in cell_text
        assert "administracion" in cell_text

    print("HIGHLIGHT_EXCEL_CONTROL_CHARACTERS_OK")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
